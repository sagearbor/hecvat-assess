#!/usr/bin/env python3
"""Generate a filled HECVAT xlsx report from assessment results JSON.

Usage:
    python3 generate_report.py <template_xlsx> <assessment_json> <output_xlsx>

Reads the assessment results (JSON with question answers and evidence),
copies the template xlsx, and fills in Answer (col C) and Additional Information
(col D) for each question across all response sheets.
"""

import json
import sys
import os
import copy
from datetime import datetime


def find_question_cells(ws, max_row):
    """Build a map of question_id -> row_number for a worksheet."""
    qmap = {}
    for row_num in range(1, max_row + 1):
        cell_a = ws.cell(row=row_num, column=1).value
        if cell_a and isinstance(cell_a, str) and "-" in cell_a:
            # Check if it looks like a HECVAT question ID (PREFIX-NN)
            parts = cell_a.strip().split("-")
            if len(parts) == 2 and parts[0].isalpha() and parts[1].isdigit():
                qmap[cell_a.strip()] = row_num
    return qmap


def generate_report(template_path: str, assessment_path: str, output_path: str):
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl required. Install with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    # Load assessment results
    with open(assessment_path) as f:
        assessment = json.load(f)

    # answers is a dict: question_id -> {answer, evidence, additional_info}
    answers = assessment.get("answers", {})

    # Load template (preserve formatting)
    wb = openpyxl.load_workbook(template_path)

    # Response sheets where answers go
    response_sheets = [
        "START HERE", "Organization", "Product", "Infrastructure",
        "IT Accessibility", "Case-Specific", "AI", "Privacy"
    ]

    filled_count = 0
    for sheet_name in response_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        qmap = find_question_cells(ws, ws.max_row)

        for qid, row_num in qmap.items():
            if qid in answers:
                ans = answers[qid]
                # Column C = Answer
                answer_val = ans.get("answer", "")
                if answer_val:
                    ws.cell(row=row_num, column=3, value=answer_val)

                # Column D = Additional Information (evidence + notes)
                additional = ans.get("additional_info", "")
                evidence = ans.get("evidence", "")
                if evidence and additional:
                    full_info = f"{additional}\n\nEvidence: {evidence}"
                elif evidence:
                    full_info = f"Evidence: {evidence}"
                elif additional:
                    full_info = additional
                else:
                    full_info = ""

                if full_info:
                    ws.cell(row=row_num, column=4, value=full_info)

                filled_count += 1

    # Update the date completed on START HERE
    if "START HERE" in wb.sheetnames:
        ws = wb["START HERE"]
        # Row 3 typically has "Date Completed" in col A, value in col C
        for row_num in range(1, 10):
            if ws.cell(row=row_num, column=1).value and "Date Completed" in str(ws.cell(row=row_num, column=1).value):
                ws.cell(row=row_num, column=3, value=datetime.now().strftime("%Y-%m-%d"))
                break

    # Save
    wb.save(output_path)
    print(f"Report generated: {output_path}")
    print(f"Filled {filled_count} question cells across {len(response_sheets)} sheets")


if __name__ == "__main__":
    if len(sys.argv) != 4:
        print(f"Usage: {sys.argv[0]} <template_xlsx> <assessment_json> <output_xlsx>",
              file=sys.stderr)
        sys.exit(1)
    generate_report(sys.argv[1], sys.argv[2], sys.argv[3])
