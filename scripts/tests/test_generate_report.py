"""Comprehensive test suite for generate_report.py

Tests cover:
- Output file creation and template preservation
- Answer filling in correct columns (C = answer, D = additional info)
- Date completion on START HERE sheet
- Edge cases: empty assessments, invalid IDs, missing data
- Sheet coverage: all response sheets processed
- No template corruption: output is valid xlsx

NOTE: Requires openpyxl for xlsx manipulation. Install with: pip install openpyxl
"""

import json
import sys
import os
from pathlib import Path
import pytest

# Add scripts directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))
from generate_report import generate_report, find_question_cells

try:
    import openpyxl
except ImportError:
    pytest.skip("openpyxl not installed", allow_module_level=True)


class TestGenerateReportFileCreation:
    """Tests validating output file creation and basic functionality.

    These tests ensure the report generator creates valid output files without
    errors or corruption.
    """

    def test_output_file_is_created(self, hecvat_xlsx_path, sample_assessment_data, tmp_path):
        """Verify output xlsx file is created.

        WHY: If file creation fails, downstream users won't have a report to
        submit. This is the most basic requirement.
        """
        assessment_file = tmp_path / "assessment.json"
        output_file = tmp_path / "output.xlsx"

        with open(assessment_file, "w") as f:
            json.dump(sample_assessment_data, f)

        assert not output_file.exists(), "Output file already exists"

        generate_report(str(hecvat_xlsx_path), str(assessment_file), str(output_file))

        assert output_file.exists(), "Output file was not created"
        assert output_file.stat().st_size > 0, "Output file is empty"

    def test_output_file_is_valid_xlsx(self, hecvat_xlsx_path, sample_assessment_data, tmp_path):
        """Verify output file can be loaded by openpyxl without errors.

        WHY: Corrupted xlsx files would be unusable. This catches binary
        corruption, invalid XML, or structural damage to the xlsx format.
        """
        assessment_file = tmp_path / "assessment.json"
        output_file = tmp_path / "output.xlsx"

        with open(assessment_file, "w") as f:
            json.dump(sample_assessment_data, f)

        generate_report(str(hecvat_xlsx_path), str(assessment_file), str(output_file))

        # Try to load the output file
        try:
            wb = openpyxl.load_workbook(str(output_file))
            assert wb is not None
            wb.close()
        except Exception as e:
            pytest.fail(f"Output xlsx is corrupted or invalid: {e}")

    def test_template_sheets_are_preserved(self, hecvat_xlsx_path, sample_assessment_data, tmp_path):
        """Verify all template sheets exist in the output file.

        WHY: Missing sheets would make the report incomplete or unusable.
        This catches bugs where sheets are accidentally deleted during processing.
        """
        assessment_file = tmp_path / "assessment.json"
        output_file = tmp_path / "output.xlsx"

        with open(assessment_file, "w") as f:
            json.dump(sample_assessment_data, f)

        # Get original sheet names
        wb_template = openpyxl.load_workbook(str(hecvat_xlsx_path))
        template_sheets = set(wb_template.sheetnames)
        wb_template.close()

        generate_report(str(hecvat_xlsx_path), str(assessment_file), str(output_file))

        # Get output sheet names
        wb_output = openpyxl.load_workbook(str(output_file))
        output_sheets = set(wb_output.sheetnames)
        wb_output.close()

        assert template_sheets == output_sheets, \
            f"Sheet mismatch. Missing: {template_sheets - output_sheets}, " \
            f"Extra: {output_sheets - template_sheets}"


class TestGenerateReportAnswerFilling:
    """Tests validating that answers are written to correct cells.

    These tests ensure the core functionality: filling Column C with answers
    and Column D with additional information.
    """

    def test_answer_fills_column_c(self, hecvat_xlsx_path, sample_assessment_data, tmp_path):
        """Verify answers are written to Column C.

        WHY: Column C is the designated answer column in HECVAT. Wrong column
        would make the report invalid and confusing to reviewers.
        """
        assessment_file = tmp_path / "assessment.json"
        output_file = tmp_path / "output.xlsx"

        with open(assessment_file, "w") as f:
            json.dump(sample_assessment_data, f)

        generate_report(str(hecvat_xlsx_path), str(assessment_file), str(output_file))

        wb = openpyxl.load_workbook(str(output_file))

        # Check GNRL-01 on START HERE sheet (should always be there)
        ws = wb["START HERE"]
        qmap = find_question_cells(ws, ws.max_row)

        if "GNRL-01" in qmap:
            row = qmap["GNRL-01"]
            answer_cell = ws.cell(row=row, column=3).value  # Column C
            assert answer_cell == "Test Vendor Inc", \
                f"GNRL-01 answer not in Column C. Got: {answer_cell}"

        wb.close()

    def test_additional_info_fills_column_d(self, hecvat_xlsx_path, sample_assessment_data, tmp_path):
        """Verify additional_info is written to Column D.

        WHY: Column D is for supplementary information. This field provides
        context and evidence for answers.
        """
        assessment_file = tmp_path / "assessment.json"
        output_file = tmp_path / "output.xlsx"

        with open(assessment_file, "w") as f:
            json.dump(sample_assessment_data, f)

        generate_report(str(hecvat_xlsx_path), str(assessment_file), str(output_file))

        wb = openpyxl.load_workbook(str(output_file))

        # Check GNRL-02 which has additional_info
        ws = wb["START HERE"]
        qmap = find_question_cells(ws, ws.max_row)

        if "GNRL-02" in qmap:
            row = qmap["GNRL-02"]
            additional_cell = ws.cell(row=row, column=4).value  # Column D
            assert additional_cell == "Cloud-based SaaS solution", \
                f"GNRL-02 additional_info not in Column D. Got: {additional_cell}"

        wb.close()

    def test_evidence_fills_column_d(self, hecvat_xlsx_path, sample_assessment_data, tmp_path):
        """Verify evidence is written to Column D with 'Evidence:' prefix.

        WHY: Evidence helps reviewers verify answers. The prefix distinguishes
        evidence from regular additional_info.
        """
        assessment_file = tmp_path / "assessment.json"
        output_file = tmp_path / "output.xlsx"

        with open(assessment_file, "w") as f:
            json.dump(sample_assessment_data, f)

        generate_report(str(hecvat_xlsx_path), str(assessment_file), str(output_file))

        wb = openpyxl.load_workbook(str(output_file))

        # AAAI-02 has evidence but no additional_info
        # Need to find which sheet AAAI-02 is on
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            qmap = find_question_cells(ws, ws.max_row)
            if "AAAI-02" in qmap:
                row = qmap["AAAI-02"]
                additional_cell = ws.cell(row=row, column=4).value
                assert additional_cell is not None, "AAAI-02 evidence not written"
                assert "Evidence:" in additional_cell, \
                    f"Evidence should have 'Evidence:' prefix. Got: {additional_cell}"
                assert "bcrypt" in additional_cell, \
                    f"Evidence content missing. Got: {additional_cell}"
                break

        wb.close()

    def test_both_additional_info_and_evidence_combined(self, hecvat_xlsx_path, sample_assessment_data, tmp_path):
        """Verify additional_info and evidence are combined in Column D.

        WHY: When both fields are present, they should be formatted together
        with proper separation. Tests the formatting logic.
        """
        assessment_file = tmp_path / "assessment.json"
        output_file = tmp_path / "output.xlsx"

        with open(assessment_file, "w") as f:
            json.dump(sample_assessment_data, f)

        generate_report(str(hecvat_xlsx_path), str(assessment_file), str(output_file))

        wb = openpyxl.load_workbook(str(output_file))

        # AAAI-01 has both additional_info and evidence
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            qmap = find_question_cells(ws, ws.max_row)
            if "AAAI-01" in qmap:
                row = qmap["AAAI-01"]
                col_d = ws.cell(row=row, column=4).value
                assert col_d is not None, "Column D is empty"
                assert "Multi-factor authentication" in col_d, \
                    f"additional_info missing. Got: {col_d}"
                assert "Evidence:" in col_d, \
                    f"Evidence prefix missing. Got: {col_d}"
                assert "auth/mfa.py" in col_d, \
                    f"Evidence content missing. Got: {col_d}"
                break

        wb.close()

    def test_multiple_sheets_are_filled(self, hecvat_xlsx_path, sample_assessment_data, tmp_path):
        """Verify answers appear on the correct sheets.

        WHY: Questions may appear on multiple sheets. All occurrences should
        be filled, not just the first one found.
        """
        assessment_file = tmp_path / "assessment.json"
        output_file = tmp_path / "output.xlsx"

        with open(assessment_file, "w") as f:
            json.dump(sample_assessment_data, f)

        generate_report(str(hecvat_xlsx_path), str(assessment_file), str(output_file))

        wb = openpyxl.load_workbook(str(output_file))

        # GNRL-01 appears on all sheets - check multiple
        gnrl_01_count = 0
        for sheet_name in ["START HERE", "Organization", "Product"]:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                qmap = find_question_cells(ws, ws.max_row)
                if "GNRL-01" in qmap:
                    row = qmap["GNRL-01"]
                    answer = ws.cell(row=row, column=3).value
                    if answer == "Test Vendor Inc":
                        gnrl_01_count += 1

        assert gnrl_01_count >= 2, \
            f"GNRL-01 should appear on multiple sheets. Found on {gnrl_01_count} sheets."

        wb.close()


class TestGenerateReportDateCompletion:
    """Tests validating date completion on START HERE sheet.

    The date completed field is important for tracking when the assessment
    was finalized.
    """

    def test_date_completed_is_filled(self, hecvat_xlsx_path, sample_assessment_data, tmp_path):
        """Verify 'Date Completed' field is filled with current date.

        WHY: Date tracking is required for audit purposes. Missing dates would
        make it unclear when the assessment was completed.
        """
        assessment_file = tmp_path / "assessment.json"
        output_file = tmp_path / "output.xlsx"

        with open(assessment_file, "w") as f:
            json.dump(sample_assessment_data, f)

        generate_report(str(hecvat_xlsx_path), str(assessment_file), str(output_file))

        wb = openpyxl.load_workbook(str(output_file))
        ws = wb["START HERE"]

        # Find "Date Completed" row
        date_found = False
        for row_num in range(1, 10):
            cell_a = ws.cell(row=row_num, column=1).value
            if cell_a and "Date Completed" in str(cell_a):
                cell_c = ws.cell(row=row_num, column=3).value
                assert cell_c is not None, "Date Completed (Column C) is empty"
                # Should be in YYYY-MM-DD format
                assert "-" in str(cell_c), f"Date format looks wrong: {cell_c}"
                date_found = True
                break

        assert date_found, "Could not find 'Date Completed' field on START HERE sheet"
        wb.close()

    def test_date_is_current_date(self, hecvat_xlsx_path, sample_assessment_data, tmp_path):
        """Verify date completed matches today's date.

        WHY: Stale or incorrect dates would be misleading. The date should
        reflect when the report was generated.
        """
        from datetime import datetime

        assessment_file = tmp_path / "assessment.json"
        output_file = tmp_path / "output.xlsx"

        with open(assessment_file, "w") as f:
            json.dump(sample_assessment_data, f)

        generate_report(str(hecvat_xlsx_path), str(assessment_file), str(output_file))

        wb = openpyxl.load_workbook(str(output_file))
        ws = wb["START HERE"]

        expected_date = datetime.now().strftime("%Y-%m-%d")

        for row_num in range(1, 10):
            cell_a = ws.cell(row=row_num, column=1).value
            if cell_a and "Date Completed" in str(cell_a):
                cell_c = ws.cell(row=row_num, column=3).value
                assert str(cell_c) == expected_date, \
                    f"Date should be {expected_date}, got {cell_c}"
                break

        wb.close()


class TestGenerateReportEdgeCases:
    """Tests for edge cases and error handling.

    These tests ensure the generator handles unusual inputs gracefully without
    crashing or corrupting the output.
    """

    def test_empty_assessment_produces_valid_file(self, hecvat_xlsx_path, empty_assessment_data, tmp_path):
        """Verify empty assessment (no answers) produces valid xlsx.

        WHY: Users may generate a blank report template for manual completion.
        Empty data should not crash or corrupt the file.
        """
        assessment_file = tmp_path / "assessment.json"
        output_file = tmp_path / "output.xlsx"

        with open(assessment_file, "w") as f:
            json.dump(empty_assessment_data, f)

        # Should not raise exception
        generate_report(str(hecvat_xlsx_path), str(assessment_file), str(output_file))

        assert output_file.exists(), "Output file not created for empty assessment"

        # Verify it's valid
        wb = openpyxl.load_workbook(str(output_file))
        assert wb is not None
        wb.close()

    def test_invalid_question_ids_are_skipped(self, hecvat_xlsx_path, assessment_with_invalid_ids, tmp_path):
        """Verify questions with non-existent IDs are silently skipped.

        WHY: Assessment may contain IDs from different HECVAT versions or
        custom questions. Invalid IDs should not crash the generator.
        """
        assessment_file = tmp_path / "assessment.json"
        output_file = tmp_path / "output.xlsx"

        with open(assessment_file, "w") as f:
            json.dump(assessment_with_invalid_ids, f)

        # Should not raise exception
        generate_report(str(hecvat_xlsx_path), str(assessment_file), str(output_file))

        assert output_file.exists(), "Output file not created"

        # Verify valid question was filled
        wb = openpyxl.load_workbook(str(output_file))
        ws = wb["START HERE"]
        qmap = find_question_cells(ws, ws.max_row)

        if "GNRL-01" in qmap:
            row = qmap["GNRL-01"]
            answer = ws.cell(row=row, column=3).value
            assert answer == "Valid Answer", "Valid question was not filled"

        wb.close()

    def test_missing_answer_field_handled_gracefully(self, hecvat_xlsx_path, tmp_path):
        """Verify missing 'answer' key doesn't crash.

        WHY: Malformed assessment data could have missing fields. The generator
        should handle this gracefully rather than crashing with KeyError.
        """
        malformed_data = {
            "answers": {
                "GNRL-01": {
                    # Missing "answer" field
                    "additional_info": "Some info"
                }
            }
        }

        assessment_file = tmp_path / "assessment.json"
        output_file = tmp_path / "output.xlsx"

        with open(assessment_file, "w") as f:
            json.dump(malformed_data, f)

        # Should not raise exception
        generate_report(str(hecvat_xlsx_path), str(assessment_file), str(output_file))

        assert output_file.exists()

    def test_none_values_handled_gracefully(self, hecvat_xlsx_path, tmp_path):
        """Verify None values in answer/additional_info/evidence are handled.

        WHY: JSON may contain null values. These should be treated as empty
        strings, not cause TypeErrors or write "None" to cells.
        """
        data_with_nones = {
            "answers": {
                "GNRL-01": {
                    "answer": "Valid",
                    "additional_info": None,
                    "evidence": None
                },
                "GNRL-02": {
                    "answer": None,
                    "additional_info": "Info",
                    "evidence": None
                }
            }
        }

        assessment_file = tmp_path / "assessment.json"
        output_file = tmp_path / "output.xlsx"

        with open(assessment_file, "w") as f:
            json.dump(data_with_nones, f)

        # Should not raise exception
        generate_report(str(hecvat_xlsx_path), str(assessment_file), str(output_file))

        wb = openpyxl.load_workbook(str(output_file))
        ws = wb["START HERE"]
        qmap = find_question_cells(ws, ws.max_row)

        # Verify None values don't appear as string "None"
        for row_num in range(1, ws.max_row + 1):
            for col_num in [3, 4]:  # Columns C and D
                cell_value = ws.cell(row=row_num, column=col_num).value
                if cell_value:
                    assert "None" != str(cell_value), \
                        f"Cell ({row_num}, {col_num}) contains string 'None'"

        wb.close()


class TestGenerateReportSheetCoverage:
    """Tests verifying all expected sheets are processed.

    These tests ensure the generator doesn't skip sheets or fail to find
    questions on certain sheets.
    """

    def test_all_response_sheets_are_processed(self, hecvat_xlsx_path, sample_assessment_data, tmp_path):
        """Verify all response sheets can be processed without errors.

        WHY: If sheet processing fails for any sheet, questions on that sheet
        would be unfilled. This catches sheet-specific errors.
        """
        assessment_file = tmp_path / "assessment.json"
        output_file = tmp_path / "output.xlsx"

        with open(assessment_file, "w") as f:
            json.dump(sample_assessment_data, f)

        # Should process all sheets without exception
        generate_report(str(hecvat_xlsx_path), str(assessment_file), str(output_file))

        wb = openpyxl.load_workbook(str(output_file))

        expected_sheets = [
            "START HERE", "Organization", "Product", "Infrastructure",
            "IT Accessibility", "Case-Specific", "AI", "Privacy"
        ]

        for sheet_name in expected_sheets:
            assert sheet_name in wb.sheetnames, \
                f"Expected sheet '{sheet_name}' not found in output"

        wb.close()

    def test_find_question_cells_returns_dict(self, hecvat_xlsx_path):
        """Verify find_question_cells helper function returns proper dict.

        WHY: This helper is critical for locating questions. If it fails or
        returns wrong data type, all question filling would fail.
        """
        wb = openpyxl.load_workbook(str(hecvat_xlsx_path))
        ws = wb["START HERE"]

        qmap = find_question_cells(ws, ws.max_row)

        assert isinstance(qmap, dict), f"Expected dict, got {type(qmap)}"
        assert len(qmap) > 0, "Question map is empty"

        # Verify structure: keys are question IDs, values are row numbers
        for qid, row_num in qmap.items():
            assert isinstance(qid, str), f"Question ID should be string, got {type(qid)}"
            assert isinstance(row_num, int), f"Row number should be int, got {type(row_num)}"
            assert "-" in qid, f"Question ID should contain hyphen: {qid}"

        wb.close()

    def test_at_least_one_question_filled_per_sheet(self, hecvat_xlsx_path, sample_assessment_data, tmp_path):
        """Verify at least one question is filled on major sheets.

        WHY: If no questions are filled on a sheet, it indicates the question
        lookup or sheet processing failed for that sheet.
        """
        assessment_file = tmp_path / "assessment.json"
        output_file = tmp_path / "output.xlsx"

        with open(assessment_file, "w") as f:
            json.dump(sample_assessment_data, f)

        generate_report(str(hecvat_xlsx_path), str(assessment_file), str(output_file))

        wb = openpyxl.load_workbook(str(output_file))

        # Check that START HERE has at least one filled answer
        ws = wb["START HERE"]
        filled_cells = 0
        for row_num in range(1, ws.max_row + 1):
            cell_c = ws.cell(row=row_num, column=3).value
            if cell_c and str(cell_c).strip():
                filled_cells += 1

        assert filled_cells > 0, "No answers filled on START HERE sheet"

        wb.close()


class TestFindQuestionCellsFunction:
    """Tests for the find_question_cells helper function.

    This function is critical for question lookup - it maps question IDs to
    row numbers. Bugs here would break all answer filling.
    """

    def test_identifies_valid_question_ids(self, hecvat_xlsx_path):
        """Verify function correctly identifies HECVAT question ID format.

        WHY: Question IDs must match pattern PREFIX-NN (e.g., AAAI-01).
        False positives would include non-question cells; false negatives
        would skip valid questions.
        """
        wb = openpyxl.load_workbook(str(hecvat_xlsx_path))
        ws = wb["START HERE"]

        qmap = find_question_cells(ws, ws.max_row)

        # All keys should match HECVAT pattern
        for qid in qmap.keys():
            parts = qid.split("-")
            assert len(parts) == 2, f"Invalid ID format: {qid}"
            assert parts[0].isalpha(), f"Prefix should be alphabetic: {qid}"
            assert parts[1].isdigit(), f"Suffix should be numeric: {qid}"

        wb.close()

    def test_maps_to_correct_row_numbers(self, hecvat_xlsx_path):
        """Verify row numbers in map point to cells containing the question IDs.

        WHY: If row numbers are off by one or incorrect, answers would be
        written to wrong rows, corrupting the report.
        """
        wb = openpyxl.load_workbook(str(hecvat_xlsx_path))
        ws = wb["START HERE"]

        qmap = find_question_cells(ws, ws.max_row)

        # Verify a few mappings are correct
        for qid, row_num in list(qmap.items())[:5]:
            cell_value = ws.cell(row=row_num, column=1).value
            assert cell_value is not None, f"Row {row_num} Column A is empty"
            # Clean whitespace for comparison
            assert qid in str(cell_value).strip(), \
                f"Row {row_num} should contain '{qid}', got '{cell_value}'"

        wb.close()

    def test_handles_empty_worksheet(self):
        """Verify function returns empty dict for empty worksheet.

        WHY: Edge case handling - function should not crash on empty/invalid
        input, but return an empty result.
        """
        # Create a minimal empty workbook
        wb = openpyxl.Workbook()
        ws = wb.active

        qmap = find_question_cells(ws, ws.max_row)

        assert isinstance(qmap, dict)
        assert len(qmap) == 0, "Empty worksheet should produce empty question map"

        wb.close()
