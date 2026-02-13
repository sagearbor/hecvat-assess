#!/usr/bin/env python3
"""Parse HECVAT xlsx into structured JSON for fast reference.

Usage:
    python3 parse_hecvat.py <xlsx_path> <output_json_path>

Extracts all questions from the 'Questions' sheet into a flat JSON array.
Each entry includes: id, question, category, compliant_response, default_importance,
default_weight, sheet_presence (which sheets the question appears on),
score_mapping, guidance fields, and repo_assessable flag.
"""

import json
import sys
import os
from datetime import datetime

def parse_hecvat(xlsx_path: str, output_path: str) -> dict:
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl required. Install with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Questions"]

    # Header row is row 2
    headers = [c.value for c in ws[2]]

    # Sheet presence columns (index 2-9): Start, Org, Product, Infra, Access, Case, AI, Privacy
    sheet_names = ["START HERE", "Organization", "Product", "Infrastructure",
                   "IT Accessibility", "Case-Specific", "AI", "Privacy"]

    # Repo-assessable question prefixes/patterns
    # These categories contain questions that can typically be answered from code
    REPO_ASSESSABLE_PREFIXES = {
        "AAAI",  # Authentication, Authorization, Account Management
        "APPL",  # Application/Service Security
        "CHNG",  # Change Management (partially)
        "DATA",  # Data (partially - encryption, handling)
        "VULN",  # Vulnerability Management
        "ITAC",  # IT Accessibility (partially - WCAG checks)
        "AIML",  # AI/ML Model Security
        "AILM",  # AI Language Models
        "AISC",  # AI Security Controls
        "AIGN",  # AI Governance (partially)
        "AIQU",  # AI Quality
        "DPAI",  # Data Privacy - AI
    }

    # Specific question IDs that are repo-assessable even if prefix isn't
    REPO_ASSESSABLE_IDS = {
        "DOCU-05",  # Architecture diagrams (check for docs)
        "DCTR-01", "DCTR-02", "DCTR-03",  # Data center (IaC/cloud config)
        "OPEM-01", "OPEM-02",  # Operational management (monitoring config)
        "DRPV-01", "DRPV-02",  # DR/Privacy (backup config)
        "CONS-01", "CONS-02",  # Consulting (dependency config)
        "THRD-01", "THRD-02", "THRD-03",  # Third party (dependency manifests)
        "PCID-01", "PCID-02",  # PCI (payment code patterns)
        "HIPA-01", "HIPA-02", "HIPA-03",  # HIPAA (PHI handling patterns)
        "FIDP-01", "FIDP-02", "FIDP-03",  # FERPA/data privacy
        "PDAT-01", "PDAT-02", "PDAT-03", "PDAT-04",  # Privacy data
        "PPPR-01", "PPPR-02", "PPPR-03",  # Privacy practices
    }

    # IDs that are never repo-assessable (organizational attestation only)
    NEVER_ASSESSABLE_PREFIXES = {
        "GNRL",  # General info (company name, contacts)
        "COMP",  # Company info (staff, operations)
        "REQU",  # Required questions (product type routing)
        "PCOM",  # Privacy company info
        "PRGN",  # Privacy regional
        "INTL",  # International
        "HFIH",  # Health/Financial/Insurance
        "PCHG",  # Privacy change management
        "PTHP",  # Privacy third party
    }

    questions = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        qid = row[0]
        if not qid or not isinstance(qid, str) or "-" not in qid:
            continue

        prefix = qid.split("-")[0]

        # Determine repo-assessability
        if prefix in NEVER_ASSESSABLE_PREFIXES:
            repo_assessable = False
        elif prefix in REPO_ASSESSABLE_PREFIXES or qid in REPO_ASSESSABLE_IDS:
            repo_assessable = True
        else:
            # Default: not repo-assessable (organizational attestation)
            repo_assessable = False

        # Sheet presence
        sheets_present = []
        for i, sname in enumerate(sheet_names):
            val = row[2 + i] if (2 + i) < len(row) else None
            if val and val == 1:
                sheets_present.append(sname)

        question = {
            "id": qid,
            "category": prefix,
            "question": row[1],
            "sheets": sheets_present,
            "score_mapping": row[10],
            "score_location": row[11],
            "has_additional_info": bool(row[12]),
            "if_then": row[13] if row[13] else None,
            "standard_guidance": row[14] if row[14] else None,
            "no_guidance": row[15] if row[15] else None,
            "yes_guidance": row[16] if row[16] else None,
            "na_guidance": row[17] if row[17] else None,
            "reason": row[18] if row[18] else None,
            "followup": row[19] if row[19] else None,
            "compliant_response": row[20] if row[20] else None,
            "default_importance": row[22] if row[22] else None,
            "default_weight": row[23] if row[23] else None,
            "repo_assessable": repo_assessable,
        }
        questions.append(question)

    # Build the output
    result = {
        "source_file": os.path.basename(xlsx_path),
        "parsed_at": datetime.utcnow().isoformat() + "Z",
        "version": "4.1.4",
        "total_questions": len(questions),
        "repo_assessable_count": sum(1 for q in questions if q["repo_assessable"]),
        "org_attestation_count": sum(1 for q in questions if not q["repo_assessable"]),
        "categories": sorted(set(q["category"] for q in questions)),
        "questions": questions,
    }

    with open(output_path, "w") as f:
        json.dump(result, f, indent=2)

    print(f"Parsed {result['total_questions']} questions "
          f"({result['repo_assessable_count']} repo-assessable, "
          f"{result['org_attestation_count']} org-attestation)")
    print(f"Output: {output_path}")
    return result


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(f"Usage: {sys.argv[0]} <xlsx_path> <output_json_path>", file=sys.stderr)
        sys.exit(1)
    parse_hecvat(sys.argv[1], sys.argv[2])
